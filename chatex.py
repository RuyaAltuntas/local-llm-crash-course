from typing import List, Dict, Any
import os
from pathlib import Path

import chainlit as cl
from groq import AsyncGroq
from dotenv import load_dotenv

from PyPDF2 import PdfReader
from docx import Document
import openpyxl

# Load variables from .env file automatically
load_dotenv()

print("Initializing Groq client...")

api_key = os.getenv("GROQ_API_KEY")
if not api_key:
    raise RuntimeError(
        "GROQ_API_KEY is not set. Put it in a .env file like:\n"
        "GROQ_API_KEY=gsk_...\n"
    )

client = AsyncGroq(api_key=api_key)

print("Client initialized.")


def extract_text_from_file(file_path: str) -> str:
    ext = Path(file_path).suffix.lower()
    try:
        if ext == '.pdf':
            reader = PdfReader(file_path)
            text = ''
            for page in reader.pages:
                text += page.extract_text() + '\n'
            return text
        elif ext == '.docx':
            doc = Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
            return text
        elif ext == '.xlsx':
            wb = openpyxl.load_workbook(file_path, data_only=True)
            text = ''
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f'Sheet: {sheet_name}\n'
                for row in sheet.iter_rows(values_only=True):
                    row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                    text += row_text + '\n'
                text += '\n'
            return text
        else:
            return f"Unsupported file type: {ext}"
    except Exception as e:
        return f"Error extracting text: {str(e)}"


def build_messages(instruction: str, history: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    system = (
        "You are a friendly, helpful AI assistant for relatives and loved ones of military veterans.\n"
        "You answer general questions about veterans' mental and physical health, well-being, and daily life.\n"
        "You do NOT have access to any specific veteran's medical records or personal information.\n\n"
        "Guidelines:\n"
        "- Focus on supporting the relative: their concerns, feelings, and questions.\n"
        "- Give clear, practical suggestions on how to support and communicate with a veteran.\n"
        "- You may explain general concepts from health reports or medical/psychological terms in simple language.\n"
        "- Always stay general: do NOT interpret or diagnose a specific person's condition or test results.\n"
        "- Do NOT give medical advice, diagnoses, or treatment recommendations.\n"
        "- Encourage talking to qualified healthcare professionals for any specific medical or crisis concerns.\n"
        "- If the user mentions self-harm, suicide, or immediate danger, gently encourage them to contact "
        "emergency services or a crisis hotline right away, instead of trying to handle it yourself.\n\n"
        "Keep responses clear, empathetic, and focused. Aim for at most a few short paragraphs, "
        "but don’t cut useful explanations short.\n\n"
        "Think step by step before providing your final answer."
    )

    messages: List[Dict[str, str]] = [{"role": "system", "content": system}]
    messages.extend(history)
    messages.append({"role": "user", "content": instruction})
    return messages


@cl.on_chat_start
async def on_chat_start():
    cl.user_session.set("message_history", [])

    # Optional: show welcome markdown on first load
    try:
        welcome_md = Path("chainlit.md").read_text(encoding="utf-8")
        await cl.Message(content=welcome_md).send()
    except FileNotFoundError:
        await cl.Message(
            content="Welcome to ValorCare AI. You can ask general questions about supporting a veteran."
        ).send()


@cl.on_message
async def on_message(message: cl.Message):
    history: List[Dict[str, str]] = cl.user_session.get("message_history") or []

    settings = cl.user_session.get("settings", {})
    hide_chain_of_thought = settings.get("hideChainOfThought", False)

    user_content = message.content
    if message.elements:
        for element in message.elements:
            if hasattr(element, 'path') and element.path:
                extracted = extract_text_from_file(element.path)
                user_content += f"\n\n--- Content from {element.name} ---\n{extracted}\n--- End of {element.name} ---\n"

    msg = cl.Message(content="")
    await msg.send()

    messages = build_messages(user_content, history)

    response_text = ""

    try:
        stream = await client.chat.completions.create(
            model="llama-3.1-8b-instant",  # Groq model
            messages=messages,
            max_tokens=800,
            temperature=0.7,
            stream=True,
        )
    except Exception:
        await msg.update(
            content=(
                "I’m having trouble reaching the AI service right now. "
                "Please try again in a little while."
            )
        )
        return

    async for chunk in stream:
        delta = chunk.choices[0].delta.content or ""
        if not delta:
            continue
        await msg.stream_token(delta)
        response_text += delta

    # Process the response based on settings
    if hide_chain_of_thought:
        # Try to extract the final answer
        if "Final answer:" in response_text:
            final_answer = response_text.split("Final answer:", 1)[1].strip()
            await msg.update(content=final_answer)
        else:
            # If no "Final answer:", assume the whole response is the answer
            await msg.update()
    else:
        await msg.update()

    history.append({"role": "user", "content": message.content})
    history.append({"role": "assistant", "content": response_text})

    MAX_TURNS = 4
    if len(history) > MAX_TURNS * 2:
        history = history[-MAX_TURNS * 2:]

    cl.user_session.set("message_history", history)
